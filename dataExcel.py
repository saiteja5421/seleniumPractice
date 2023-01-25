import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

    @staticmethod
    def getTestData():
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\malip\\OneDrive\\Desktop\\python\\sheet1.xlsx")
        sheet = book.active
        print(sheet.max_row)
        print(sheet.max_column)
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == "hi":

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty

                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        print([Dict])

    getTestData()
